import tkinter as tk
import sqlite3 as sq
import openpyxl as op
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from sqlite3 import Error
from openpyxl import load_workbook

if __name__ == '__main__':
    main()




class Excel:
    db_name = 'vacation_shedule.db'

    def __init__(self, window):

        self.wind = window
        # создание элементов для ввода слов и значений
        frame = LabelFrame(self.wind, text='Выберите первую неделю для расчета графика')
        frame.grid(row=0, column=0, columnspan=3, pady=20)
        Label(frame, text='Мин_неделя: ').grid(row=1, column=0)
        con = sq.connect('vacation_shedule.db')
        cursor_em = con.cursor()
        # получаем все данные из таблицы calendar
        cursor_em.execute('SELECT distinct(week_number) FROM calendar ORDER BY month')
        lis = cursor_em.fetchall()
        # print(lis)
        self.combobox = ttk.Combobox(frame, values=lis)
        self.combobox.grid(row=1, column=1)

        Label(frame, text='Макс_неделя: ').grid(row=2, column=0)
        cursor_em0 = con.cursor()
        # получаем все данные из таблицы calendar
        cursor_em0.execute('SELECT distinct week_number FROM calendar ORDER BY month')
        lis0 = cursor_em0.fetchall()
        # print(lis)
        self.combobox0 = ttk.Combobox(frame, values=lis0)
        self.combobox0.grid(row=2, column=1)

        Label(frame, text='Год: ').grid(row=3, column=0)
        cursor_ye = con.cursor()
        # получаем все данные из таблицы calendar
        cursor_ye.execute('SELECT distinct year_id FROM calendar ORDER BY year_id')
        lis2 = cursor_ye.fetchall()
        self.combobox1 = ttk.Combobox(frame, values=lis2)
        self.combobox1.grid(row=3, column=1)

        b0 = ttk.Button(frame, text='Расчет прогноза исследований', command=self.pz_type_study).grid(row=4, columnspan=3, sticky=W + E)
        self.message = Label(text='', fg='green')
        self.message.grid(row=5, column=0, columnspan=3, sticky=W + E)
        b1 = ttk.Button(frame, text='Сохранить', command=self.save).grid(row=6, columnspan=3, sticky=W + E)
        # таблица прогнозных значений исследований
        columns = ("#1", "#2")
        self.tree = ttk.Treeview(frame, height=10, columns=columns)
        self.tree.grid(row=7, column=0, columnspan=3)
        self.tree.heading('#0', text='Тип исследования', anchor=CENTER)
        self.tree.heading('#1', text='Номер недели', anchor=CENTER)
        self.tree.heading('#2', text='Количество исследований', anchor=CENTER)



    # подключение и запрос к базе
    def run_query(self, query, parameters=()):
        with sq.connect(self.db_name) as conn:
            cursor = conn.cursor()
            result = cursor.execute(query, parameters)
            conn.commit()
        return result

    def validation(self):
        return len(self.combobox.get()) != 0 and len(self.combobox0.get()) != 0 and len(self.combobox1.get()) != 0

    #расчет прогнозируемых исследований
    def pz_type_study(self):
        if self.validation():
            parameters = (self.combobox.get(), self.combobox0.get())
            con = sq.connect('vacation_shedule.db')
            cursor_pz = con.cursor()
            # получаем src_id последнего расчета
            self.predraschet()
            query = '''INSERT INTO pz_vol_study
                        select b4.src_id, b1.year_id+1, b3.week_number, b3.id_type_study, b2.avg_12
                        from
                            (select dy.id_type_study, dy.year_id, dy.avg_dov_year
                            , lag(dy.avg_dov_year) over (partition by dy.id_type_study order by dy.year_id) as avg_dov_year_2022
                            from 	
                                /*Средняя по году с учетом доверительных исследований*/
                                (select ag.year_id, ag.id_type_study, ag.avg_year, avg(ag.VolPcs) avg_dov_year
                                from
                                    /*исключаем отклонение от средней более чем 50%*/
                                    (select vl.year_id, vl.week_number, vl.id_type_study, cast(vl.VolPcs as integer) as VolPcs
                                    , cast(ay.avg_year as integer) as avg_year
                                    , case when cast(vl.VolPcs as integer)<ay.avg_year*3/4 then 0 else 1 end as pr_dov_week
                                    from volpcs_study vl
                                    join 
                                    /*Средняя по году без исключений недель*/
                                        (select year_id, id_type_study, cast(avg(volPcs) as integer) as avg_year
                                        from volpcs_study
                                        where VolPcs is not null
                                        and year_id>2021
                                        and year_id<2024
                                        GROUP by year_id, id_type_study) ay on vl.year_id = ay.year_id and vl.id_type_study = ay.id_type_study
                                    where vl.VolPcs is not null) ag
                                where ag.pr_dov_week = 1
                                group by ag.year_id, ag.id_type_study, ag.avg_year) dy) b1
                        join 
                        /*Средняя последних 12 доверительных недель*/
                        (select a1.id_type_study, avg(a1.VolPcs) as avg_12
                        from
                            (select vd.year_id, vd.week_number, vd.id_type_study, vd.VolPcs
                            , row_number() over (partition by vd.year_id, vd.id_type_study order by vd.week_number desc) as rw
                            from volpcs_study vd
                            join
                                    /*исключаем отклонение от средней более чем 50%*/
                                    (select vl.year_id, vl.week_number, vl.id_type_study, cast(vl.VolPcs as integer) as VolPcs
                                    , cast(ay.avg_year as integer) as avg_year
                                    , case when cast(vl.VolPcs as integer)<ay.avg_year*3/4 then 0 else 1 end as pr_dov_week
                                    from volpcs_study vl
                                    join 
                                    /*Средняя по году без исключений недель*/
                                        (select year_id, id_type_study, cast(avg(volPcs) as integer) as avg_year
                                        from volpcs_study
                                        where VolPcs is not null
                                        and year_id>2021
                                        and year_id<2024
                                        GROUP by year_id, id_type_study) ay on vl.year_id = ay.year_id and vl.id_type_study = ay.id_type_study
                                    where vl.VolPcs is not null) ag on vd.year_id=ag.year_id and vd.week_number=ag.week_number and vd.id_type_study=ag.id_type_study
                                    and vd.year_id = 2023
                                    and ag.pr_dov_week = 1) a1 
                            where a1.rw<=12
                            group by a1.id_type_study) b2 on b1.id_type_study = b2.id_type_study
                        join volpcs_study b3 on b1.id_type_study=b3.id_type_study and b1.year_id=b3.year_id
                        cross join (select max(src_id) as src_id from tab_employees_days) b4
                        where b1.year_id = 2023
                        and b3.week_number >=?
                        and b3.week_number<=?
                        group by b4.src_id, b1.year_id+1, b3.week_number, b3.id_type_study, b2.avg_12'''
            pz_rows = self.run_query(query, parameters)
            self.message['text'] = 'Расчет произведен'
            self.get_study_pz()
            self.get_shedule_main()
            self.get_shedule_main_1()



        else:
            self.message['text'] = 'Введите год и месяц для прогноза исследований'


    # собираем данные для расчета по сотрудникам с учетом отпусков, больничных и нерабочих дней
    def predraschet(self):
        con = sq.connect('vacation_shedule.db')
        parameters = (self.combobox1.get(), self.combobox.get(), self.combobox0.get())
        query = '''INSERT INTO tab_employees_days
                    select b4.src_id+1 as src_id, a1.date, a1.week_number, a1.year_id
                    , a1.id_employees, a1.id_type_study, a1.bid, a1.bid_hours, nd.date_with, nd.date_to
                    , case when ed.id_employees is null then 0 else 1 end as pr_dop_study, NULL as work_hours_study
                    from
                        (select date, week_number, year_id, id_employees, id_type_study, bid, bid_hours
                        from
                            (select es.id_employees, es.id_type_study, em.bid, em.bid*12 as bid_hours
                            from employees_study es
                            join employees1 em on es.id_employees = em.id)
                        cross join 
                            (select *
                            from calendar ca
                            where year_id=? and week_number>=? and week_number<=? and ca.is_work is null)) a1
                    left join not_working_days nd on a1.id_employees=nd.id_employees
                    left join (select distinct id_employees from employees_study_dop) ed on a1.id_employees=ed.id_employees
                    cross join (select max(src_id) as src_id from tab_employees_days) b4
                    where nd.id_employees is NULL
                    or a1.date<nd.date_with or a1.date>nd.date_to'''
        db_rows = self.run_query(query,parameters)


    # заполнение таблицы прогнозируемым количеством исследований
    def get_study_pz(self):
        con = sq.connect('vacation_shedule.db')
        cursor_sh = con.cursor()
        # получаем src_id последнего расчета
        cursor_sh.execute('SELECT max(distinct(src_id)) FROM tab_employees_days')
        max_src_id = cursor_sh.fetchall()
        max_src_id = max_src_id[0]  # list to tuple
        max_src_id = max_src_id[0]  # tuple to char
        max_src_id = int(max_src_id)  # char to int
        #print('max_src_id=',max_src_id)
        records = self.tree.get_children()
        for element in records:
            self.tree.delete(element)
        parameters = (self.combobox1.get(), self.combobox.get(), self.combobox0.get(), max_src_id)
        con = sq.connect('vacation_shedule.db')
        cursor_pz_vw= con.cursor()
        # получаем src_id последнего расчета
        query='''SELECT ts.name, vp.week_number, vp.vol_pcs 
                    FROM pz_vol_study vp join type_of_study ts on vp.id_type_study=ts.id
                    where vp.year_id=? and vp.week_number>=? and vp.week_number<=? and vp.src_id=? 
                    ORDER BY ts.name, vp.week_number DESC'''
        db_rows = self.run_query(query,parameters).fetchall()

        for row in db_rows:
            self.tree.insert('', 0, text=row[0], values=row[1:])


    # Расчет графика по основным исследованиям с учетом отпусков, больничных и праздников(уже собранных)
    def get_shedule_main(self):
        con = sq.connect('vacation_shedule.db')
        cursor_sh = con.cursor()
        # получаем src_id последнего расчета
        cursor_sh.execute('SELECT max(distinct(src_id)) FROM tab_employees_days')
        max_src = cursor_sh.fetchall()
        max_src = max_src[0]  # list to tuple
        max_src = max_src[0]  # tuple to char
        max_src = int(max_src)  # char to int
        # получаем список основных исследований, по которому будем запускать цикл
        cursor_st = con.cursor()
        cursor_st.execute(
            'SELECT distinct(id_type_study) FROM tab_employees_days where src_id = ? ORDER BY id_type_study',
            str(max_src))
        study_rows = cursor_st.fetchall()
        # print(study_rows)
        lim1 = len(study_rows)
        list_study = list()
        list_employees = list()
        # print(lim1)
        # начинаем цикл расчета по исследованиям из списка
        for i in range(0, lim1):
            n = study_rows[i]
            # print(n)
            n1 = int(n[0])
            # print(n1)
            list_study.insert(i, n1)
            #определяем список дней для расчета
            parameters = (max_src, n1)
            query = '''select distinct(date_calendar)
                        from tab_employees_days
                        where src_id = ?
                        and id_type_study = ?
                        order by week_number asc'''
            days_rows = self.run_query(query, parameters).fetchall()
            for k in range(0, len(days_rows)):
                parameters_pz_days = (max_src, max_src, (days_rows[k])[0], n1)

                query_pz_days = '''select (cast(a1.vol_pcs/7 as int) + 1) - (coalesce(a3.sum_work_hours, 0)*(cast((a4.max_norma+a4.min_norma)/2/8 as int)+1)) as pz_days_per
                                    from pz_vol_study a1
                                    join calendar a2 on a1.week_number = a2.week_number
                                    join (select src_id, date_calendar, id_type_study, sum(work_hours_study) sum_work_hours
                                        from tab_employees_days
                                        where src_id=?
                                        group by date_calendar, id_type_study) a3 on a2.date = a3.date_calendar and a1.id_type_study = a3.id_type_study
                                    join norma_rc a4 on a1.id_type_study = a4.id_type_study
                                    where a1.src_id = ?
                                    and a2.date = ?
                                    and a1.id_type_study = ?'''
                pz_days = self.run_query(query_pz_days, parameters_pz_days).fetchall()

                #выбираем сотрудников для дальнейшего цикла
                parameters_em = (max_src, (days_rows[k])[0], n1)
                query_em = '''select distinct(id_employees)
                                    from tab_employees_days em
                                    join calendar ca on em.date_calendar
                                    where em.src_id = ?
                                    and em.date_calendar = ?
                                    and em.id_type_study = ?
                                    and em.work_hours_study is null
                                    and pr_dop_study=0
                                    order by id_employees'''
                em_rows = self.run_query(query_em, parameters_em).fetchall()

                #производим расчет по дням
                for j in range(0, len(em_rows)):
                    # считаем рабочие дни
                    parameters_wd = (max_src, max_src, (days_rows[k])[0], n1, (em_rows[j])[0], max_src, (em_rows[j])[0], (days_rows[k])[0], n1)
                    #print('em_rows0', (em_rows[j])[0])
                    query_wd = '''UPDATE tab_employees_days
                            SET work_hours_study = 
                                (select case when emw.sum_week<40*em.bid then 
                                    case when 40*em.bid-emw.sum_week <12 then 40*em.bid-emw.sum_week else 12 end 
                                    else 0 end as wok_h
                                from tab_employees_days em
                                join calendar ca on em.date_calendar=ca.date
                                left join (select id_employees, ca.week_number, sum(coalesce(em.work_hours_study, 0)) as sum_week
                                    from tab_employees_days em
                                    join calendar ca on em.date_calendar=ca.date
                                    where src_id = ?
                                    group by id_employees, ca.week_number
                                    ) emw on em.id_employees=emw.id_employees and emw.week_number=ca.week_number
                                where em.src_id = ?
                                and em.date_calendar = ?
                                and em.id_type_study = ?
                                and em.work_hours_study is null
                                and em.pr_dop_study=0
                                and em.id_employees = ?
                                )
                            where src_id = ?
                            and id_employees = ?
                            and date_calendar = ?
                            and id_type_study = ?'''
                    wd_rows = self.run_query(query_wd, parameters_wd).fetchall()

                    # пересчитываем оставшийся ПЗ для расчпределения специалистов
                    parameters_pzd_p = (max_src, max_src, (days_rows[k])[0], n1)

                    query_pzd_p = '''select (cast(a1.vol_pcs/7 as int) + 1) - (coalesce(a3.sum_work_hours, 0)*(cast((a4.max_norma+a4.min_norma)/2/8 as int)+1)) as pz_days_per
                                                                            from pz_vol_study a1
                                                                            join calendar a2 on a1.week_number = a2.week_number
                                                                            join (select src_id, date_calendar, id_type_study, sum(work_hours_study) sum_work_hours
                                                                                from tab_employees_days
                                                                                where src_id=?
                                                                                group by date_calendar, id_type_study) a3 on a2.date = a3.date_calendar and a1.id_type_study = a3.id_type_study
                                                                            join norma_rc a4 on a1.id_type_study = a4.id_type_study
                                                                            where a1.src_id = ?
                                                                            and a2.date = ?
                                                                            and a1.id_type_study = ?'''
                    pz_days_p = self.run_query(query_pzd_p, parameters_pzd_p).fetchall()
                    #print('pz_days_p[0]', (pz_days_p[0])[0])
                    if (pz_days_p[0])[0] <=0:
                        break


# Расчет графика по основным исследованиям с учетом отпусков, больничных и праздников(уже собранных)
    def get_shedule_main_1(self):
        con = sq.connect('vacation_shedule.db')
        cursor_sh = con.cursor()
        # получаем src_id последнего расчета
        cursor_sh.execute('SELECT max(distinct(src_id)) FROM tab_employees_days')
        max_src = cursor_sh.fetchall()
        max_src = max_src[0]  # list to tuple
        max_src = max_src[0]  # tuple to char
        max_src = int(max_src)  # char to int
        # получаем список основных исследований, по которому будем запускать цикл
        cursor_st = con.cursor()
        cursor_st.execute(
            'SELECT distinct(id_type_study) FROM tab_employees_days where src_id = ? ORDER BY id_type_study',
            str(max_src))
        study_rows = cursor_st.fetchall()
        # print(study_rows)
        lim1 = len(study_rows)
        list_study = list()
        list_employees = list()
        # print(lim1)
        # начинаем цикл расчета по исследованиям из списка
        for i in range(0, lim1):
            n = study_rows[i]
            # print(n)
            n1 = int(n[0])
            # print(n1)
            list_study.insert(i, n1)
            #определяем список дней для расчета
            parameters = (max_src, n1)
            query = '''select distinct(date_calendar)
                        from tab_employees_days
                        where src_id = ?
                        and id_type_study = ?
                        order by week_number asc'''
            days_rows = self.run_query(query, parameters).fetchall()
            for k in range(0, len(days_rows)):
                parameters_pz_days = (max_src, max_src, (days_rows[k])[0], n1)

                query_pz_days = '''select (cast(a1.vol_pcs/7 as int) + 1) - (coalesce(a3.sum_work_hours, 0)*(cast((a4.max_norma+a4.min_norma)/2/8 as int)+1)) as pz_days_per
                                    from pz_vol_study a1
                                    join calendar a2 on a1.week_number = a2.week_number
                                    join (select src_id, date_calendar, id_type_study, sum(work_hours_study) sum_work_hours
                                        from tab_employees_days
                                        where src_id=?
                                        group by date_calendar, id_type_study) a3 on a2.date = a3.date_calendar and a1.id_type_study = a3.id_type_study
                                    join norma_rc a4 on a1.id_type_study = a4.id_type_study
                                    where a1.src_id = ?
                                    and a2.date = ?
                                    and a1.id_type_study = ?'''
                pz_days = self.run_query(query_pz_days, parameters_pz_days).fetchall()

                #выбираем сотрудников для дальнейшего цикла
                parameters_em = (max_src, (days_rows[k])[0], n1)
                query_em = '''select distinct(id_employees)
                                    from tab_employees_days em
                                    join calendar ca on em.date_calendar
                                    where em.src_id = ?
                                    and em.date_calendar = ?
                                    and em.id_type_study = ?
                                    and em.work_hours_study is null
                                    and pr_dop_study=1
                                    order by id_employees'''
                em_rows = self.run_query(query_em, parameters_em).fetchall()

                #производим расчет по дням
                for j in range(0, len(em_rows)):
                    # считаем рабочие дни
                    parameters_wd = (max_src, max_src, (days_rows[k])[0], n1, (em_rows[j])[0], max_src, (em_rows[j])[0], (days_rows[k])[0], n1)
                    #print('em_rows0', (em_rows[j])[0])
                    query_wd = '''UPDATE tab_employees_days
                            SET work_hours_study = 
                                (select case when emw.sum_week<40*em.bid then 
                                    case when 40*em.bid-emw.sum_week <12 then 40*em.bid-emw.sum_week else 12 end 
                                    else 0 end as wok_h
                                from tab_employees_days em
                                join calendar ca on em.date_calendar=ca.date
                                left join (select id_employees, ca.week_number, sum(coalesce(em.work_hours_study, 0)) as sum_week
                                    from tab_employees_days em
                                    join calendar ca on em.date_calendar=ca.date
                                    where src_id = ?
                                    group by id_employees, ca.week_number
                                    ) emw on em.id_employees=emw.id_employees and emw.week_number=ca.week_number
                                where em.src_id = ?
                                and em.date_calendar = ?
                                and em.id_type_study = ?
                                and em.work_hours_study is null
                                and em.pr_dop_study=1
                                and em.id_employees = ?
                                )
                            where src_id = ?
                            and id_employees = ?
                            and date_calendar = ?
                            and id_type_study = ?'''
                    wd_rows = self.run_query(query_wd, parameters_wd).fetchall()

                    # пересчитываем оставшийся ПЗ для расчпределения специалистов
                    parameters_pzd_p = (max_src, max_src, (days_rows[k])[0], n1)

                    query_pzd_p = '''select (cast(a1.vol_pcs/7 as int) + 1) - (coalesce(a3.sum_work_hours, 0)*(cast((a4.max_norma+a4.min_norma)/2/8 as int)+1)) as pz_days_per
                                                                            from pz_vol_study a1
                                                                            join calendar a2 on a1.week_number = a2.week_number
                                                                            join (select src_id, date_calendar, id_type_study, sum(work_hours_study) sum_work_hours
                                                                                from tab_employees_days
                                                                                where src_id=?
                                                                                group by date_calendar, id_type_study) a3 on a2.date = a3.date_calendar and a1.id_type_study = a3.id_type_study
                                                                            join norma_rc a4 on a1.id_type_study = a4.id_type_study
                                                                            where a1.src_id = ?
                                                                            and a2.date = ?
                                                                            and a1.id_type_study = ?'''
                    pz_days_p = self.run_query(query_pzd_p, parameters_pzd_p).fetchall()
                    #print('pz_days_p[0]', (pz_days_p[0])[0])
                    if (pz_days_p[0])[0] <=0:
                        break

    def save(self):
        # создаем книгу
        wb = op.Workbook()
        # делаем единственный лист активным
        ws = wb.active
        ws.title = "График работы сотрудников"
        # Меняем цвет названия книги
        ws.sheet_properties.tabColor = "1072BA"
        con = sq.connect('vacation_shedule.db')
        cursor_sh = con.cursor()
        # получаем src_id последнего расчета
        cursor_sh.execute('SELECT max(distinct(src_id)) FROM tab_employees_days')
        max_src = cursor_sh.fetchall()
        max_src = max_src[0]  # list to tuple
        max_src = max_src[0]  # tuple to char
        max_src = int(max_src)  # char to int
        parameters_ex = (str(max_src))
        print('max_src',max_src)
        query_ex = '''select em.src_id, ep.name, ts.name, em.date_calendar, em.date_with, em.date_to, em.bid, em.work_hours_study
                        from tab_employees_days em
                        join type_of_study ts on em.id_type_study = ts.id
                        join employees1 ep on em.id_employees = ep.id
                        where src_id = ?'''
        row_ex = self.run_query(query_ex, parameters_ex).fetchall()
        #print(row_ex)
        query_ex_count = '''select count(*)
                                from tab_employees_days em
                                join type_of_study ts on em.id_type_study = ts.id
                                join employees1 ep on em.id_employees = ep.id
                                where src_id = ?'''
        row_ex_count = self.run_query(query_ex_count, parameters_ex).fetchall()
        ws["A1"] = "Номер расчета"
        ws["B1"] = "ФИО"
        ws["C1"] = "Тип исследования"
        ws["D1"] = "Дата"
        ws["E1"] = "Дата начала отпуска/больничного"
        ws["F1"] = "Дата окончания отпуска/больничного"
        ws["G1"] = "Ставка"
        ws["H1"] = "Смена"
        for h in range(2, (row_ex_count[0])[0]):
            for g in range(0, 8):
                ws.cell(row=h, column=g+1).value = (row_ex[h])[g]
        wb.save('test.xlsx')




if __name__ == '__main__':
    window = Tk()
    application = Excel(window)
    window.mainloop()



